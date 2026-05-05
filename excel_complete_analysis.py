import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

def analyze_excel_complete(file_path, range_a1_z30="A1:Z30"):
    """Vollständige Analyse der Excel-Datei im Bereich A1:Z30"""
    print(f"=== Vollständige Excel-Analyse: {file_path} ===")
    print(f"Bereich: {range_a1_z30}")
    print()
    
    # Mit pandas alle Daten lesen
    df = pd.read_excel(file_path, sheet_name='Projektwürdigkeitsanalyse', header=None)
    
    # Mit openpyxl für Formeln und detaillierte Zell-Infos
    wb = load_workbook(file_path)
    ws = wb['Projektwürdigkeitsanalyse']
    
    print("=== GRAFIKÜBERSICHT A1:Z30 ===")
    for row in range(1, 31):  # Zeilen 1-30
        row_data = []
        for col in range(1, 27):  # Spalten A-Z (1-26)
            cell_addr = f"{get_column_letter(col)}{row}"
            cell = ws[cell_addr]
            
            if cell.value is not None:
                if cell.data_type == 'f':  # Formel
                    row_data.append(f"FORMEL:{cell.value}")
                else:
                    row_data.append(str(cell.value)[:20])  # Begrenzte Länge für Übersicht
            else:
                row_data.append("")
        
        # Nur Zeilen mit Inhalt anzeigen
        if any(row_data):
            print(f"Zeile {row:2d}: {' | '.join(row_data[:10])}...")  # Erste 10 Spalten
    
    print("\n=== DETAILLIERTE ZELLANALYSE ===")
    
    # Wichtige Bereiche identifizieren
    important_cells = {}
    
    # Header-Bereich
    print("\n--- HEADER-BEREICH ---")
    for row in range(1, 7):
        for col in range(1, 27):
            cell_addr = f"{get_column_letter(col)}{row}"
            cell = ws[cell_addr]
            if cell.value is not None:
                print(f"{cell_addr}: {cell.value}")
                if cell.data_type == 'f':
                    print(f"  Formel: {cell.value}")
                important_cells[cell_addr] = cell.value
    
    # Kriterien-Bereich (Zeilen 7-17)
    print("\n--- KRITERIEN-BEREICH (Zeilen 7-17) ---")
    criteria_data = {}
    
    for row in range(7, 18):  # Zeilen 7-17
        criterion_info = {
            'criterion': None,
            'change_option': None,
            'kleinprojekt_option': None,
            'projekt_option': None,
            'change_points': None,
            'kleinprojekt_points': None,
            'projekt_points': None,
            'selection': None
        }
        
        # Kriterium (Spalte B)
        criterion_cell = ws[f'B{row}']
        if criterion_cell.value:
            criterion_info['criterion'] = criterion_cell.value
            
            # Change-Option (Spalte D)
            change_opt_cell = ws[f'D{row}']
            if change_opt_cell.value:
                criterion_info['change_option'] = change_opt_cell.value
            
            # Kleinprojekt-Option (Spalte F)
            kleinprojekt_opt_cell = ws[f'F{row}']
            if kleinprojekt_opt_cell.value:
                criterion_info['kleinprojekt_option'] = kleinprojekt_opt_cell.value
            
            # Projekt-Option (Spalte H)
            projekt_opt_cell = ws[f'H{row}']
            if projekt_opt_cell.value:
                criterion_info['projekt_option'] = projekt_opt_cell.value
            
            # Punkte (Spalten P, Q, R)
            change_points_cell = ws[f'P{row}']
            if change_points_cell.value:
                criterion_info['change_points'] = change_points_cell.value
            
            kleinprojekt_points_cell = ws[f'Q{row}']
            if kleinprojekt_points_cell.value:
                criterion_info['kleinprojekt_points'] = kleinprojekt_points_cell.value
            
            projekt_points_cell = ws[f'R{row}']
            if projekt_points_cell.value:
                criterion_info['projekt_points'] = projekt_points_cell.value
            
            # Auswahl (Spalte J)
            selection_cell = ws[f'J{row}']
            if selection_cell.value:
                criterion_info['selection'] = selection_cell.value
            
            criteria_data[row] = criterion_info
            print(f"\nZeile {row}:")
            print(f"  Kriterium: {criterion_info['criterion']}")
            print(f"  Change: {criterion_info['change_option']} ({criterion_info['change_points']} Punkte)")
            print(f"  Kleinprojekt: {criterion_info['kleinprojekt_option']} ({criterion_info['kleinprojekt_points']} Punkte)")
            print(f"  Projekt: {criterion_info['projekt_option']} ({criterion_info['projekt_points']} Punkte)")
            print(f"  Auswahl: {criterion_info['selection']}")
    
    # Ergebnis-Bereich
    print("\n--- ERGEBNIS-BEREICH ---")
    
    # Gesamt-Punkte (Zeile 18)
    print("Zeile 18 - Gesamt-Punkte:")
    for col in ['E', 'G', 'I']:
        cell = ws[f'{col}18']
        if cell.value is not None:
            print(f"  {col}18: {cell.value}")
            if cell.data_type == 'f':
                print(f"    Formel: {cell.value}")
    
    # Empfehlung (Zeile 19)
    print("\nZeile 19 - Empfehlung:")
    for col in range(1, 27):
        cell_addr = f"{get_column_letter(col)}19"
        cell = ws[cell_addr]
        if cell.value is not None:
            print(f"  {cell_addr}: {cell.value}")
            if cell.data_type == 'f':
                print(f"    Formel: {cell.value}")
    
    # Formeln analysieren
    print("\n=== ALLE FORMELN IM BEREICH A1:Z30 ===")
    formulas = {}
    
    for row in range(1, 31):
        for col in range(1, 27):
            cell_addr = f"{get_column_letter(col)}{row}"
            cell = ws[cell_addr]
            if cell.data_type == 'f':  # Formel
                formulas[cell_addr] = cell.value
                print(f"{cell_addr}: {cell.value}")
    
    # Labels und Konstanten
    print("\n=== LABELS UND KONSTANTEN ===")
    
    # Wichtige Labels finden
    labels = {
        'change_label': None,
        'kleinprojekt_label': None,
        'projekt_label': None
    }
    
    # D5, F5, H5 sollten die Labels enthalten
    d5 = ws['D5'].value
    f5 = ws['F5'].value
    h5 = ws['H5'].value
    
    labels['change_label'] = d5
    labels['kleinprojekt_label'] = f5
    labels['projekt_label'] = h5
    
    print(f"Change Label (D5): {labels['change_label']}")
    print(f"Kleinprojekt Label (F5): {labels['kleinprojekt_label']}")
    print(f"Projekt Label (H5): {labels['projekt_label']}")
    
    return {
        'criteria_data': criteria_data,
        'formulas': formulas,
        'labels': labels,
        'important_cells': important_cells
    }

def generate_streamlit_data(excel_data):
    """Generiere Datenstruktur für Streamlit-App"""
    print("\n=== STREAMLIT-DATENSTRUKTUR ===")
    
    criteria_data = {}
    for row, info in excel_data['criteria_data'].items():
        if info['criterion']:
            criteria_data[info['criterion']] = {
                'options': [
                    info['change_option'],
                    info['kleinprojekt_option'],
                    info['projekt_option']
                ],
                'points': {
                    'Change': info['change_points'],
                    'Kleinprojekt': info['kleinprojekt_points'],
                    'Projekt': info['projekt_points']
                }
            }
    
    print("Kriterien-Daten für Streamlit:")
    for criterion, data in criteria_data.items():
        print(f"  {criterion}:")
        print(f"    Options: {data['options']}")
        print(f"    Points: {data['points']}")
    
    return criteria_data

if __name__ == "__main__":
    excel_data = analyze_excel_complete("Projektwürdigkeitsanalyse_Jovo.xlsx")
    streamlit_data = generate_streamlit_data(excel_data)
